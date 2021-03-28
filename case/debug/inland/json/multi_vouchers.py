# coding=utf-8
import os
import random
import time
import simplejson
from openpyxl.reader.excel import load_workbook 
from six import with_metaclass
from lib.common.utils.meta import WithLogger
from lib.common_biz.find_key import GetKey, is_get_key_from_db
from lib.common.file_operation.config_operation import Config
from lib.common_biz.sign import Sign
from lib.common.algorithm.md5 import md5
from lib.common.utils.globals import GlobalVar
from lib.common_biz.file_path import key_path
from lib.common.case_processor.parser import Parser
from lib.common.utils.misc_utils import create_random_str


class MultiVouInfoGenerator(with_metaclass(WithLogger, Parser)):
    
    def __init__(self, path, ws_name='grant_multi_vouchers'):
        wb = load_workbook(path, data_only=True) #显示公式计算结果
        self.ws = wb[ws_name]
        self.fields = []
        self._is_case_started = False
        self.vou_info = []
    
    def parse(self, to_iter=False):
        for row in self.ws.iter_rows(self.ws.min_row, self.ws.max_row,
                                     self.ws.min_column, self.ws.max_column,
                                     values_only=False):
            # row is a tuple
            if row[0].value == 'vouType':
                [self.fields.append(cell.value)
                 for cell in row 
                 if cell.value.isascii()]
                self._is_case_started = True
                self.logger.info('字段：%s' %self.fields)
                continue            
            if self._is_case_started:
                info = [cell.value for cell in row][:-1]
#                 if to_iter:
#                     yield info
                self.vou_info.append(info)
        self.to_dict()
#         self.update_vou_info()
        return True
    
    def __iter__(self):
        result = yield from self.parse()
        if result:
            print('parse finished.')
            
    def to_dict(self):
        for idx, info in enumerate(self.vou_info):
            self.vou_info[idx] = dict(zip(self.fields, info))
    
    def update_vou_info(self):
        '''
        vouName: 如果为空，则生成32位随机字符串
        amount: 取表格中数据，或者自增一个浮点小数
        grantCount: 为0时，会从vou_info中remove
        '''
        self.vou_info = [d for d in self.vou_info if d['grantCount']]
        for d in self.vou_info:
            d['vouType'] = str(d['vouType'])
            if not d['vouName']:
                d['vouName'] = create_random_str(32)
            d['amount'] = random.choice([d['amount'], d['amount']+round(random.random(), 2)])
            if d['maxAmount'] is None:
                d['maxAmount'] = 0
            if d['ratio'] is None:
                d['ratio'] = 0
            d['beginTime'] = int(time.time() * 10**3)
            d['expireTime'] = int(time.time() * 10**3)
            d['scopeId'] = '7104f7bc23e445daba913a5a96a264ac'
            d['subScopeId'] = '7104f7bc23e445daba913a5a96a264ac'
            d['blackScopeId'] = '7104f7bc23e445daba913a5a96a264ac'
            d['batchId'] = create_random_str(32)
    
    def total(self):
        return sum([d['grantCount'] for d in self.vou_info], 0)
            

def grant_multi_vouchers(ssoid, partner_id, vouinfos:list):
    '''
    按照用户维度，批量发不同类型的券
    :param ssoid: 账户id
    :param partner_id: 业务线id
    :请求参数列表：
    字段          类型        含义        必要性
    ssoid       String      用户ID           Y
    country     String    二位字母的国家代码   N
    timezone    String    UTC时区            N
    currency    String    三位字母的币种      N
    appId       String    业务线ID           Y    
    requestId   String    请求ID            Y
    sign        String    签名              Y
    grantVoucherInfoList    List<GrantVoucherInfo>    Y
        vouType    String    券优惠类型                  Y
        vouName    String    优惠券名称                Y
        grantCount    Integer    发放数量              Y
        amount    BigDecimal    金额,元    
        maxAmount    BigDecimal    金额,元    
        ratio    BigDecimal    打折的折扣    打折券使用
        beginTime    Long    允许使用的开始时间         Y
        expireTime    Long    允许使用的结束时间        Y
        scopeId    String    使用范围ID                Y
        subScopeId    String    子范围ID              N
        blackScopeId    String    黑名单范围ID         N
        settleType    String    结算类型               Y
        batchId    String    批次号                    N
    '''
    req = {
        'ssoid': ssoid,
        'country': 'CN',
        'timezone': 'GMT+08:00',
        'currency': 'CNY',
        'appId': partner_id,
        'requestId': create_random_str(32),
        'sign': '',
        'grantVoucherInfoList': vouinfos
    }
    if is_get_key_from_db:
        priv_key = GetKey(req['appId']).get_key_from_voucher()
    else:
        priv_key = Config(key_path).as_dict('oversea_vou_app_info')["key_" + req['appId']]
    req_ = req.copy()
    req_['grantVoucherInfoList'] = simplejson.dumps(req['grantVoucherInfoList'], ensure_ascii=False).replace(' ', '')
    orig_sign_str = Sign(req_).join_asc_have_key("&key=") + priv_key
    req['sign'] = md5(orig_sign_str, to_upper=True)
    result = GlobalVar.HTTPJSON_IN.post("/voucher/grantMultiVoucher", data=req)

