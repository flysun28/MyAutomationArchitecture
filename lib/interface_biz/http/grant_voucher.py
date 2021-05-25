#!/usr/bin/env Python3
# -*- encoding:utf-8 *-*
# author:xy
# datetime:2021/3/9 23:42
# comment: 单张优惠券申请
import datetime
import random
import simplejson
import threading
from six import with_metaclass
from itertools import product
from lib.common.algorithm.md5 import md5
from lib.common.file_operation.config_operation import Config
from lib.common.utils.globals import GlobalVar
from lib.common_biz.file_path import key_path
from lib.common_biz.find_key import is_get_key_from_db, GetKey
from lib.common_biz.order_random import RandomOrder
from lib.common_biz.sign import Sign
from lib.common.utils.misc_utils import create_random_str, to_iterable_nested
from openpyxl.reader.excel import load_workbook 
from lib.common.utils.meta import WithLogger
from lib.common.exception.http_exception import HttpJsonException
from lib.common_biz.find_database_table import SeparateDbTable
from lib.common.db_operation.mysql_operation import is_connect_mysql

end_time = str((datetime.datetime.now() + datetime.timedelta(days=365)).strftime('%Y-%m-%d %H:%M:%S'))


def grant_voucher(amount=1, count=1, vou_type=1, appId="2031"):
    """
    默认消费券
    :param appId: 
    :param vou_type: 
    :param amount: 分
    :return:
    """
    req = {
        "amount": amount,
        "appId": appId,
        "appSubName": "AUTO_TEST",
        "blackScopeId": "",
        "checkName": "TEST_ACCOUNT",
        "configId": "",
        "count": count,
        "country": "CN",
        "currency": "CNY",
        "expireTime": end_time,
        "ext1": "",
        "ext2": "",
        "maxAmount": int(amount)+1,
        "name": "AUTO_TEST",
        "partnerOrder": RandomOrder(28).business_order("AUTO"),
        "ratio": round(random.random(), 2),
        "remark": "",
        "salePrice": 0,
        "scopeId": "7104f7bc23e445daba913a5a96a264ac",
        "settleType": 1,
        "sign": "",
        "ssoid": GlobalVar.SSOID,
        "subScopeId": "",
        "timezone": "",
        # 1 消费券
        "type": vou_type,
        "useableTime": "2021-01-01 00:00:00"
    }
    key = ''
    if is_get_key_from_db:
        key = GetKey(req['appId']).get_key_from_voucher()
    else:
        key = Config(key_path).as_dict('oversea_vou_app_info')["key_" + req['appId']]
    req['sign'] = md5(Sign(req).join_asc_have_key("&key=") + key)
    result = GlobalVar.HTTPJSON_IN.post("/voucher/grantSingle", data=req)
    # 返回优惠券id
    return result['vouIdList'][0]


def grant_single_voucher(ssoid=GlobalVar.SSOID, vou_type=1, appId="2031", salt_key=''):
    if vou_type == 5:
        amount = round(random.uniform(100, 1000), 2)
    else:
        amount = round(random.uniform(1, 10), 2)
    count = random.randint(1, 10)
    max_amount = round(random.uniform(amount, amount+100), 2)
    ratio = round(random.random(), 2)
    req = {
        "amount": amount,
        "appId": appId,
        "appSubName": "AUTO_TEST",
        "blackScopeId": "",
        "checkName": "TEST_ACCOUNT",
        "configId": "",
        "count": count,
        "country": "CN",
        "currency": "CNY",
        "expireTime": end_time,
        "ext1": "",
        "ext2": "",
        "maxAmount": max_amount,
        "name": "AUTO_TEST",
        "partnerOrder": RandomOrder(28).business_order("AUTO"),
        "ratio": ratio,
        "remark": "",
        "salePrice": 0,
        "scopeId": "7104f7bc23e445daba913a5a96a264ac",
        "settleType": 1,
        "sign": "",
        "ssoid": ssoid,
        "subScopeId": "",
        "timezone": "",
        "type": vou_type,
        "useableTime": "2021-01-01 00:00:00"
    }
    if salt_key:
        key = salt_key
    elif is_get_key_from_db:
        key = GetKey(req['appId']).get_key_from_voucher()
    else:
        key = Config(key_path).as_dict('oversea_vou_app_info')["key_" + req['appId']]
    req['sign'] = md5(Sign(req).join_asc_have_key("&key=") + key)
    result = GlobalVar.HTTPJSON_IN.post("/voucher/grantSingle", data=req)
    # 返回优惠券列表
    return result['vouIdList']


class VouInfo(with_metaclass(WithLogger)):
    
    def __init__(self, path, ws_name='grant_multi_vouchers'):
        wb = load_workbook(path, data_only=True) #显示公式计算结果
        self.ws = wb[ws_name]
        self.fields = []
        self._is_case_started = False
        self.vou_info = []
    
    def parse(self, to_iter=False):
        for row in self.ws.iter_rows(self.ws.min_row, self.ws.max_row,
                                     self.ws.min_column, self.ws.max_column,
                                     values_only=False):
            # row is a tuple
            if row[0].value == 'vouType':
                [self.fields.append(cell.value)
                 for cell in row 
                 if cell.value.isascii()]
                self._is_case_started = True
                self.logger.info('字段：%s' %self.fields)
                continue            
            if self._is_case_started:
                info = [cell.value for cell in row][:-1]
                if to_iter:
                    yield info
                self.vou_info.append(info)
        self._to_dict()
        return True
    
    def __iter__(self):
        result = yield from self.parse()
        if result:
            print('parse finished.')
    
    @property
    def count(self):
        return sum([d['grantCount'] for d in self.vou_info], 0)
    
    @property
    def batchIds(self):
        return [d['configId'] for d in self.vou_info]
                
    def _to_dict(self):
        for idx, info in enumerate(self.vou_info):
            self.vou_info[idx] = dict(zip(self.fields, info))
    
    def update(self):
        '''
        vouName: 如果为空，则生成32位随机字符串
        amount: 取表格中数据，或者自增一个浮点小数
        grantCount: 为0时，会从vou_info中remove
        '''
        self.vou_info = [d for d in self.vou_info if d['grantCount']]
        print(self.vou_info)
        for d in self.vou_info:
            d['vouType'] = str(d['vouType'])
            if not d['vouName']:
                d['vouName'] = create_random_str(32)
            d['amount'] = random.choice([d['amount'], d['amount']+round(random.random(), 3)])
            if d['maxAmount'] is None:
                d['maxAmount'] = 0
            if d['ratio'] is None:
                d['ratio'] = 0
            for attr in 'amount', 'maxAmount', 'ratio':
                d[attr] = str(d[attr])
            now = datetime.datetime.now()
            d['beginTime'] = int(now.timestamp() * 10**3)
            one_year_later = now + datetime.timedelta(days=365)
            d['expireTime'] = int(one_year_later.timestamp() * 10**3)
            d['scopeId'] = '7104f7bc23e445daba913a5a96a264ac'
            d['subScopeId'] = '7104f7bc23e445daba913a5a96a264ac'
            d['blackScopeId'] = '7104f7bc23e445daba913a5a96a264ac'
            d['configId'] = create_random_str(32)
    
    def create(self):
        list(self)
        self.update()
        print('表格测试数据：')
        for info in self.vou_info:
            print(info)


class HttpGrantMultiVous(with_metaclass(WithLogger)):
    partner_private_key = {'5456925': 'cm7dld743nre523kk439rd4a2f2aw3fku1'}
    
    def __init__(self, vouinfo:VouInfo, ssoid, partner_id='5456925'):
        self.vouinfo_obj = vouinfo
        self.ssoid = ssoid
        if is_connect_mysql:
            self.vou_table_id = SeparateDbTable(self.ssoid).get_vou_table()
        self.partner_id = partner_id
        self.salt_key = ''
        self.req = {}
        if is_get_key_from_db():
            self.salt_key = GetKey(self.partner_id).get_key_from_voucher()
        else:
            self.salt_key = self.partner_private_key[self.partner_id]
            self.logger.info('查询到优惠券秘钥信息: %s' %self.salt_key)
        self.request_ids = set()
        self.lock = threading.Lock()
    
    def post(self, data=None):
        '''
        按照用户维度，批量发不同类型的券
        幂等比对key: ssoid, appId, requestId
        :param ssoid: 账户id
        :param partner_id: 业务线id
        :请求参数列表：
        字段          类型        含义        必要性
        ssoid       String      用户ID           Y
        country     String    二位字母的国家代码   N
        timezone    String    UTC时区            N
        currency    String    三位字母的币种      N
        appId       String    业务线ID           Y    
        requestId   String    请求ID            Y
        sign        String    签名              Y
        grantVoucherInfoList    List<GrantVoucherInfo>    Y
            vouType    String    券优惠类型                  Y
            vouName    String    优惠券名称                Y
            grantCount    Integer    发放数量              Y
            amount    BigDecimal    金额,元    
            maxAmount    BigDecimal    金额,元    
            ratio    BigDecimal    打折的折扣    打折券使用
            beginTime    Long    允许使用的开始时间         Y
            expireTime    Long    允许使用的结束时间        Y
            scopeId    String    使用范围ID                Y
            subScopeId    String    子范围ID              N
            blackScopeId    String    黑名单范围ID         N
            settleType    String    结算类型               Y
            configId    String    批次号                    N
        '''
        if data:
            req_id = data['requestId']
        else:
            req_id = self.init_req()
            self.make_sign(self.req)
            data = self.req
        with self.lock:
            self.request_ids.add(req_id)
        result = GlobalVar.HTTPJSON_IN.post("/voucher/grantMultiVoucher", data=data)
        self.validate_response(result)
    
    def init_req(self):
        self.req = {
            'ssoid': self.ssoid,
            'country': 'CN',
            'timezone': 'GMT+08:00',
            'currency': 'CNY',
            'appId': self.partner_id,
            'requestId': create_random_str(32),
            'sign': '',
            'grantVoucherInfoList': self.vouinfo_obj.vou_info
        }
        return self.req['requestId']
    
    def make_sign(self, req):
        req_ = req.copy()
        req_['grantVoucherInfoList'] = simplejson.dumps(req['grantVoucherInfoList'], ensure_ascii=False).replace(' ', '')
        orig_sign_str = Sign(req_).join_asc_have_key("&key=") + self.salt_key
        del req_
#         print('签名原串:', orig_sign_str)
        req['sign'] = md5(orig_sign_str, to_upper=True)
    
    def validate_response(self, resp):
        assert resp['code'] == '0000', '返回参数 code != 0000'
        assert resp['msg'] == 'success', '返回参数 msg != success'
        resp_vouinfo = resp['simpleVoucherInfoDtoList']
        assert len(resp_vouinfo) == self.vouinfo_obj.count, '返回的优惠券个数 != %d' %self.vouinfo_obj.count
        resp_batch_ids = set(d['configId'] for d in resp_vouinfo)
        for batchid in resp_batch_ids:
            assert batchid in self.vouinfo_obj.batchIds, '返回的批次号 %s 不在 %s范围内' %(batchid, self.vouinfo_obj.batchIds)        
    
    def _make_neg_data(self, data, **neg_kwargs):
        '''
        :param data: a dict or a list to be updated
        :param neg_kwargs: {k1:[v11,v12,v13...], k2:[v21,v22,v23...], ...}
        '''
        for k, v in neg_kwargs.items():
            neg_kwargs[k] = to_iterable_nested(v, ele_type=tuple)
        for param_values in product(*neg_kwargs.values()):
            req = data.copy()
            upd_kw = dict(zip(neg_kwargs.keys(), param_values))
            print('当前负向参数:', upd_kw)
            req.update(upd_kw)
            yield req
            del req
    
    def _do_neg_test(self, req):
        self.make_sign(req)
        try:
            self.post(req)
        except AssertionError as e:
            self.logger.info('异常测试期望失败-实际失败：%s' %e)
        except:
            raise
        else:
            self.logger.error('异常测试期望失败-实际成功：%s' %req)
            raise HttpJsonException('异常测试期望失败-实际成功：%s' %req)
    
    def common_params_negative_test(self, **neg_kwargs):
        '''
        :param neg_kwargs: {k1:[v11,v12,v13...], k2:[v21,v22,v23...], ...}
        '''
        self.init_req()
        sql = "SELECT max(id) as lastId FROM oppopay_voucher.vou_info_%d WHERE ssoid='%s' ORDER BY id DESC;" %(self.vou_table_id, self.ssoid)
        max_id = GlobalVar.MYSQL_IN.select_one(sql)['lastId']
        self.logger.info('负向测试开始......')
        for idx, req in enumerate(self._make_neg_data(self.req, **neg_kwargs), 1):
            if idx == 1:
                continue
            self._do_neg_test(req)
        # 查询是否产生了新记录，期望无
        sql = "SELECT * FROM oppopay_voucher.vou_info_%d WHERE ssoid='%s' AND id>%d ORDER BY id DESC;" %(self.vou_table_id, max_id, self.ssoid)
        result = GlobalVar.MYSQL_IN.select(sql)
        assert result == ()
        self.logger.info('负向测试结束......')
    
    def vouinfolist_negative_test(self, **neg_kwargs):
        self.init_req()
        req = self.req.copy()
        sql = "SELECT max(id) as lastId FROM oppopay_voucher.vou_info_%d WHERE ssoid='%s' ORDER BY id DESC;" %(self.vou_table_id, self.ssoid)
        max_id = GlobalVar.MYSQL_IN.select_one(sql)['lastId']
        self.logger.info('负向测试开始......')
        for idx, vouinfolist in enumerate(self._make_neg_data(self.req['grantVoucherInfoList'][0], **neg_kwargs), 1):
            if idx == 1:
                continue
            req['grantVoucherInfoList'][0] = vouinfolist
            self._do_neg_test(req)
        # 查询是否产生了新记录，期望无
        sql = "SELECT * FROM oppopay_voucher.vou_info_%d WHERE ssoid='%s' AND id>%d ORDER BY id DESC;" %(self.vou_table_id, max_id, self.ssoid)
        result = GlobalVar.MYSQL_IN.select(sql)
        assert result == ()
        self.logger.info('负向测试结束......')


class HttpGrantSingleVous(HttpGrantMultiVous):
    
    def __init__(self, vou_type, ssoid=GlobalVar.SSOID, partner_id="5456925"):
        self.ssoid = ssoid
        if is_connect_mysql:
            self.vou_table_id = SeparateDbTable(self.ssoid).get_vou_table()
        self.vou_type = vou_type
        self.partner_id = partner_id
        self.req = None
        if is_get_key_from_db():            
            self.salt_key = GetKey(self.partner_id).get_key_from_voucher()
        else:            
            self.salt_key = self.partner_private_key[self.partner_id]
            self.logger.info('查询到优惠券秘钥信息: %s' %self.salt_key)
    
    def init_req(self, count=None):
        if self.vou_type == 5:
            amount = round(random.uniform(100, 1000), 2)
        else:
            amount = round(random.uniform(1, 10), 2)
        count = count or random.randint(1, 10)
        max_amount = round(random.uniform(amount, amount+100), 2)
        ratio = round(random.uniform(0.01, 0.99), 2)
        end_time = str((datetime.datetime.now() + datetime.timedelta(days=365)).strftime('%Y-%m-%d %H:%M:%S'))
        self.req = {
            "amount": amount,
            "appId": self.partner_id,
            "appSubName": "AUTO_TEST",
            "blackScopeId": "",
            "checkName": "TEST_ACCOUNT",
            "configId": "",
            "count": count,
            "country": "CN",
            "currency": "CNY",
            "expireTime": end_time,
            "ext1": "",
            "ext2": "",
            "maxAmount": max_amount,
            "name": "AUTO_TEST",
            "partnerOrder": RandomOrder(28).business_order("AUTO"),
            "ratio": ratio,
            "remark": "",
            "salePrice": 0,
            "scopeId": "7104f7bc23e445daba913a5a96a264ac",
            "settleType": random.choice([0, 1]),
            "sign": "",
            "ssoid": self.ssoid,
            "subScopeId": "",
            "timezone": "",
            "type": self.vou_type,
            "useableTime": "2021-01-01 00:00:00"
        }
    
    def make_sign(self, req):
        req_ = req.copy()
        orig_sign_str = Sign(req_).join_asc_have_key("&key=") + self.salt_key
        del req_
#         print('签名原串:', orig_sign_str)
        req['sign'] = md5(orig_sign_str, to_upper=True)
    
    def post(self, count=None, data=None):
        self.init_req(count)
        self.make_sign(self.req)
        req = data if data else self.req
        result = GlobalVar.HTTPJSON_IN.post("/voucher/grantSingle", data=req)
        assert result['code'] == '0000', '返回参数 code != 0000'
        assert result['msg'] == 'success', '返回参数 msg != success'
        # 返回优惠券列表
        self.logger.info('优惠券id列表：%s' %result['vouIdList'])
        return result['vouIdList']
    

if __name__ == '__main__':
    print(grant_voucher())
    