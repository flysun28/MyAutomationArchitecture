'''
Created on 2021年2月8日
@author: 80319739
'''
import os
import re
from _functools import partial
from six import with_metaclass
from abc import ABCMeta, abstractmethod
from _collections_abc import _check_methods
from lib import pardir
from openpyxl.worksheet.worksheet import Worksheet
from lib.common.utils.meta import WithLogger
from lib.common.file_operation.excel_operation import Excel
from .case import ExcelTestCaseHttp, ExcelTestCaseDubbo
from lib.common.utils.descriptors import CoordinateDescriptor


class Parser(metaclass=ABCMeta):
    
    def __iter__(self):
        finished = yield from self.generate()
        finished = 'Yes' if finished else 'No'
        self.logger.info('Are all cases in <Worksheet "%s"> are generated? %s' %(self.ws.title, finished))
        
    @abstractmethod
    def parse(self, *args, **kwargs):
        raise NotImplementedError
    
    generate = partial(parse, to_iter=True)
    
    @classmethod
    def __subclasshook__(cls, C):
        if cls is Parser:
            return _check_methods(C, "parse")
        return NotImplemented


class ExcelParser(with_metaclass(WithLogger, Parser)):
    
    def __new__(cls, fileobj:Excel, interface):
        cls.fd_coord = {}
        cls.fields = None
        self = object.__new__(cls)
        self.sheetname = interface
        self.init()
        self.ws = fileobj.open_worksheet(interface)
        for row in self.ws.iter_rows(self.ws.min_row, self.ws.max_row,
                                    self.ws.min_column, self.ws.max_column,
                                    values_only=False):
            # row is a tuple
            if row[0].value == '用例名称':
                self.fields = tuple(cell.value for cell in row)
                for cell in row:
                    self.fd_coord.setdefault(cell.value, cell.coordinate)
        cls.case_name_coord = CoordinateDescriptor('用例名称')
        cls.actual_coord = CoordinateDescriptor('实际结果')
        cls.running_result_coord = CoordinateDescriptor('用例执行结果')
        cls.req_coord = CoordinateDescriptor('请求报文体')
        cls.voucher_coord = CoordinateDescriptor('voucherInfo')
        return self
    
    def __init__(self, fileobj:Excel, interface):
        self.protocol = re.search('http|dubbo|protobuf', fileobj.path, re.I).group()
        self.testcase_cls = eval('ExcelTestCase' + os.path.basename(pardir(fileobj.path)).capitalize())

    def init(self):        
        self._is_case_started = False
        self.cases = []

    def parse(self, to_iter=False):
        self.init()
        for row in self.ws.iter_rows(self.ws.min_row, self.ws.max_row,
                                     self.ws.min_column, self.ws.max_column,
                                     values_only=False):
            # row is a tuple
            if row[0].value.upper() == 'URL':
                self.interface_url = row[1].value
                continue
            if row[0].value == '用例名称':
                self._is_case_started = True
#                 self.fields = tuple(cell.value for cell in row)
#                 for cell in row:
#                     self.fd_coord.setdefault(cell.value, cell.coordinate)
                continue
            if self._is_case_started:
                _case = self.testcase_cls(self.fields, self.protocol)
                _case.ws = self.ws
                _case.fill(cell.value for cell in row)
                if to_iter:
                    yield _case
                else:
                    self.cases.append(_case)
        return True

