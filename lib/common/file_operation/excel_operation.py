'''
@author: 80319739
'''
# import asyncio
from openpyxl import load_workbook, Workbook
from lib.common.concurrent.threading import ResultTakenThread
from lib.common.utils.meta import WithLogger


class Excel(metaclass=WithLogger):
    workbooks = {}
    
    def __init__(self, path, read_only=False):
        self.path = path
        self.wb = None
#         self.wb = load_workbook(self.path, read_only=True)
        self.loading_thr = ResultTakenThread(load_workbook,
                                             self.path,
                                             read_only=read_only,
                                             daemon=True,
                                             name='LoadingExcelWorkbookThread')
        # 对于同一个文件，wb只能被打开一次，否则实际上对应的是不同的实例，写入时只有最后一个实例会生效
        self.wb = self.workbooks.setdefault(self.path, self.loading_thr.result)
        assert type(self.wb) is Workbook, 'invalid workbook obj %s' %self.wb
#         asyncio.run(self.loading_workbook())
#     
#     async def loading_workbook(self):
#         await load_workbook(self.path, read_only=True)
        
    def open_worksheet(self, name):
        return self.wb[name]
    
    def save(self):
        self.wb.save(self.path)
    
    def close(self):
        self.wb.close()

    def sheetnames(self):
        return self.wb.sheetnames
